# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

"""
Import KRA, KPI, and Appraisal Templates from the KRA-KPI Excel file.

Usage:
  bench --site <site> execute hrms.hr.doctype.appraisal.import_kra_kpi.import_kra_kpi_from_excel \
    --kwargs '{"file_path": "/path/to/KRA-KPI_NAsty 2026.xlsx"}'

Or from browser console (with Frappe File doc name or file_url):
  frappe.call({
    method: "hrms.hr.doctype.appraisal.import_kra_kpi.import_kra_kpi_from_excel",
    args: { file_path: "FILE-00001" },
    callback: (r) => console.log(r.message)
  })
"""

import io
import os

import frappe
import requests
from frappe import _
from frappe.utils import flt


@frappe.whitelist()
def import_kra_kpi_from_excel(file_path):
	"""Read all department sheets from the KRA-KPI Excel and create records.

	Args:
		file_path: One of:
			- Absolute local path (e.g. /home/user/file.xlsx)
			- Frappe file URL (e.g. /files/file.xlsx, /private/files/file.xlsx)
			- Frappe File doc name (e.g. FILE-00001)
			- S3 file URL (e.g. /api/method/frappe_s3_attachment.controller.generate_file?key=...)

	Returns:
		dict with counts of created records
	"""
	import openpyxl

	wb = openpyxl.load_workbook(_resolve_file(file_path), data_only=True)

	summary = {
		"kras_created": 0,
		"kras_skipped": 0,
		"kpis_created": 0,
		"kpis_skipped": 0,
		"templates_created": 0,
		"templates_skipped": 0,
		"errors": [],
	}

	for sheet_name in wb.sheetnames:
		if not sheet_name.startswith("KRA"):
			continue

		# Extract department name from sheet: "KRA·KPI Global Sales" → "Global Sales"
		dept_name = sheet_name.replace("KRA·KPI", "").replace("KRA.KPI", "").strip()
		if not dept_name:
			continue

		ws = wb[sheet_name]
		positions = _parse_sheet(ws)

		for pos in positions:
			_create_records_for_position(pos, dept_name, summary)

	frappe.db.commit()

	msg = (
		f"Import complete: {summary['kras_created']} KRAs, "
		f"{summary['kpis_created']} KPIs, "
		f"{summary['templates_created']} templates created. "
		f"{summary['kras_skipped']} KRAs, "
		f"{summary['kpis_skipped']} KPIs, "
		f"{summary['templates_skipped']} templates skipped (already exist)."
	)
	if summary["errors"]:
		msg += f" {len(summary['errors'])} errors."

	frappe.msgprint(msg)
	print(msg)
	if summary["errors"]:
		for err in summary["errors"]:
			print(f"  ERROR: {err}")

	return summary


def _resolve_file(file_path):
	"""Resolve file_path to a file-like object or local path.

	Handles: local paths, Frappe file URLs, File doc names, S3 presigned URLs.
	Returns: file path string (local) or BytesIO (remote/S3).
	"""
	# 1. Local file path
	if os.path.exists(file_path):
		return file_path

	# 2. Frappe local file URL (/files/... or /private/files/...)
	if file_path.startswith("/files/") or file_path.startswith("/private/files/"):
		local_path = frappe.get_site_path(file_path.lstrip("/"))
		if os.path.exists(local_path):
			return local_path
		# Fall through to S3 resolution

	# 3. Resolve File doc — either by name or by file_url
	file_doc = None
	if not file_path.startswith("/"):
		# Looks like a File doc name (e.g. FILE-00001)
		if frappe.db.exists("File", file_path):
			file_doc = frappe.get_doc("File", file_path)
	else:
		# Looks like a URL — find the File doc
		results = frappe.get_all("File", filters={"file_url": file_path}, limit=1)
		if results:
			file_doc = frappe.get_doc("File", results[0].name)

	if not file_doc:
		frappe.throw(_(f"File not found: {file_path}"))

	# 4. Check if file is on local filesystem (non-S3)
	if file_doc.file_url:
		local_path = frappe.get_site_path(file_doc.file_url.lstrip("/"))
		if os.path.exists(local_path):
			return local_path

	# 5. S3 file — get presigned URL and download
	return _download_s3_file(file_doc)


def _download_s3_file(file_doc):
	"""Download file from S3 via presigned URL. Returns BytesIO."""
	from frappe_s3_attachment.controller import S3Operations

	key = _extract_s3_key(file_doc.file_url)
	if not key:
		frappe.throw(_(f"Cannot extract S3 key from file URL: {file_doc.file_url}"))

	s3 = S3Operations()
	presigned_url = s3.get_url(key, file_doc.file_name)

	response = requests.get(presigned_url, timeout=60)
	response.raise_for_status()

	return io.BytesIO(response.content)


def _extract_s3_key(file_url):
	"""Extract S3 key from frappe_s3_attachment URL.

	Private files: /api/method/frappe_s3_attachment.controller.generate_file?key={key}&file_name={name}
	Public files: https://s3.endpoint.com/bucket/key (key is the path after bucket)
	"""
	if not file_url:
		return None

	from urllib.parse import parse_qs, urlparse

	parsed = urlparse(file_url)

	# Private file — key is in query string
	if "generate_file" in (parsed.path or ""):
		qs = parse_qs(parsed.query)
		keys = qs.get("key", [])
		return keys[0] if keys else None

	# Public file — key is the path after /bucket/
	# URL format: https://endpoint/bucket/key
	path_parts = parsed.path.strip("/").split("/", 1)
	if len(path_parts) > 1:
		return path_parts[1]

	return None


def _parse_sheet(ws):
	"""Parse a single worksheet into a list of position dicts.

	Each position has:
		title: str (e.g. "COUNTRY DIRECTOR  |  Grade: E1")
		grade: str (e.g. "E1")
		kpis: list of dicts with kra, kpi, param, weight, target
	"""
	positions = []
	current_pos = None
	current_kra = None

	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
		line = str(row[0] or "").strip()

		# Detect position header: "▸ COUNTRY DIRECTOR  |  Grade: E1"
		if "▸" in line or line.startswith(">"):
			if current_pos:
				positions.append(current_pos)

			title = line.replace("▸", "").replace(">", "").strip()

			# Extract grade if present
			grade = ""
			if "Grade:" in title:
				parts = title.split("Grade:")
				grade = parts[-1].strip()
				title = parts[0].strip().rstrip("|").strip()

			current_pos = {"title": title, "grade": grade, "kpis": []}
			current_kra = None
			continue

		if not current_pos:
			continue

		# Check if it's a numbered row
		try:
			int(line)
		except (ValueError, TypeError):
			continue

		kra = str(row[1] or "").strip()
		if kra:
			current_kra = kra

		kpi_title = str(row[2] or "").strip()
		param = str(row[3] or "").strip()
		weight = flt(row[4])
		target = str(row[5] or "").strip()

		if kpi_title and current_kra:
			current_pos["kpis"].append({
				"kra": current_kra,
				"kpi": kpi_title,
				"param": param,
				"weight": weight,
				"target": target,
			})

	if current_pos:
		positions.append(current_pos)

	return positions


def _create_records_for_position(pos, dept_name, summary):
	"""Create KRA, KPI, and Appraisal Template records for one position."""

	template_title = f"{dept_name} - {pos['title']}"

	# Skip if template already exists
	if frappe.db.exists("Appraisal Template", template_title):
		summary["templates_skipped"] += 1
		return

	goals = []

	for kpi_data in pos["kpis"]:
		kra_name = kpi_data["kra"]
		kpi_title = kpi_data["kpi"]

		# Create KRA if not exists
		if not frappe.db.exists("KRA", kra_name):
			try:
				frappe.get_doc({
					"doctype": "KRA",
					"title": kra_name,
					"department": dept_name if frappe.db.exists("Department", dept_name) else None,
				}).insert(ignore_permissions=True)
				summary["kras_created"] += 1
			except Exception as e:
				summary["errors"].append(f"KRA '{kra_name}': {e}")
				continue
		else:
			summary["kras_skipped"] += 1

		# Create KPI if not exists
		kpi_name = f"{kra_name}-{kpi_title}"
		if not frappe.db.exists("KPI", kpi_name):
			try:
				frappe.get_doc({
					"doctype": "KPI",
					"title": kpi_title,
					"kra": kra_name,
					"description": kpi_data["param"],
					"default_target": _parse_target(kpi_data["target"]),
				}).insert(ignore_permissions=True)
				summary["kpis_created"] += 1
			except Exception as e:
				summary["errors"].append(f"KPI '{kpi_title}' under '{kra_name}': {e}")
				continue
		else:
			summary["kpis_skipped"] += 1

		goals.append({
			"key_result_area": kra_name,
			"kpi": kpi_name,
			"kpi_description": kpi_data["param"],
			"per_weightage": kpi_data["weight"],
			"default_target": _parse_target(kpi_data["target"]),
		})

	if not goals:
		return

	# Create Appraisal Template
	try:
		template = frappe.get_doc({
			"doctype": "Appraisal Template",
			"template_title": template_title,
			"department": dept_name if frappe.db.exists("Department", dept_name) else None,
			"goals": goals,
		})
		template.insert(ignore_permissions=True)
		summary["templates_created"] += 1
	except Exception as e:
		summary["errors"].append(f"Template '{template_title}': {e}")


def _parse_target(target_str):
	"""Try to extract a numeric target from strings like '≥100% of target', '≥15 per year', etc."""
	if not target_str:
		return 0

	import re

	# Extract first number from the string
	match = re.search(r"[\d.]+", target_str)
	if match:
		return flt(match.group())
	return 0
